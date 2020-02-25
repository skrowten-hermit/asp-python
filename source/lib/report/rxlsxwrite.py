


########################################################################################################################
########################################################################################################################
## File              :: vaca_rxlsx_write_routines.py
## Description       :: VoIP Automation Common API : Spreadsheet (.xlsx) report test-case result data population.
## Developer         :: Sreekanth S
## Version           :: v1.0
## Release Date      :: 11/04/2019
## Changes made      :: Initial version.
## Changes made Date :: 11/04/2019
## Changes made by   :: Sreekanth S
########################################################################################################################
########################################################################################################################



import time
import openpyxl
import re
from copy import copy
from openpyxl.styles import Font



class eXLSXreport:
    def __init__(self, debug=0):
        self.path = ""
        self.report = ""
        self.ram = 0
        self.css = 0
        self.app = 0
        self.java = 0
        self.dfdok = 0
        self.bmr = 0
        self.rows_RUI = 1
        self.cols_RUI = 1
        self.rows_CAM = 1
        self.cols_CAM = 1
        self.rows_JSM = 1
        self.cols_JSM = 1
        self.rows_DBM = 1
        self.cols_DBM = 1
        self.rows_BMR = 1
        self.cols_BMR = 1
        self.workbook = openpyxl.Workbook()
        self.worksheet_RUI = ""
        self.worksheet_CAM = ""
        self.worksheet_JSM = ""
        self.worksheet_DBM = ""
        self.worksheet_BMR = ""
        self.cell_RUI = ""
        self.cell_CAM = ""
        self.cell_JSM = ""
        self.cell_DBM = ""
        self.cell_BMR = ""
        self.num_format = '#0.000000'
        self.bold_font = openpyxl.styles.Font(bold=True)
        self.DEBUG = debug


    def create_perfReport(self, fname, ram, css, app, java, dfdok, bmr):
        self.report = fname
        self.ram, self.css, self.app, self.java, self.dfdok, self.bmr = ram, css, app, java, dfdok, bmr
        self.workbook.save(self.report)
        if self.DEBUG == 1:
            print "Workbook created at path : ", self.report
        self.bgc = openpyxl.styles.PatternFill(start_color="00AACC00", end_color="00AACC00", fill_type='solid')

        if self.ram:
            self.worksheet_RUI = self.workbook.create_sheet()
            self.worksheet_RUI.title = 'RAM_Usage_Information'
            for sheet in self.workbook:
                if sheet.title == 'Sheet':
                    self.workbook.remove_sheet(sheet)

##  The compilation of RAM Usage Information worksheet starts here.
##  Writing the headers for the columns in RAM Usage Information worksheet

            self.cell_RUI = self.worksheet_RUI.cell(row=self.rows_RUI, column=self.cols_RUI)
            self.cell_RUI.value = self.suite_info
            self.cell_RUI.font = self.bold_font
            self.cell_RUI.fill = self.bgc
            self.rows_RUI += 1
            self.cell_RUI = self.worksheet_RUI.cell(row=self.rows_RUI, column=self.cols_RUI)
            self.cell_RUI.value = "Scenario"
            self.cell_RUI.font = self.bold_font
            self.cell_RUI.fill = self.bgc
            self.cell_RUI = self.worksheet_RUI.cell(row=self.rows_RUI, column=self.cols_RUI + 1)
            self.cell_RUI.value = "Memory Free (kB)"
            self.cell_RUI.font = self.bold_font
            self.cell_RUI.fill = self.bgc
            self.workbook.save(self.report)

            self.rows_RUI += 1

        if self.css or self.app:
            self.worksheet_CAM = self.workbook.create_sheet()
            self.worksheet_CAM.title = 'CSS-AP_MHz_Measurement'
            for sheet in self.workbook:
                if sheet.title == 'Sheet':
                    self.workbook.remove_sheet(sheet)
# For looper without DDR usage data
            if self.app == 1:
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM)
                self.cell_CAM.value = self.suite_info
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.rows_CAM += 1
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM)
                self.cell_CAM.value = "Scenario"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 1)
                self.cell_CAM.value = "Looper"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 2)
                self.cell_CAM.value = "CSS Term CPU Free %"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 3)
                self.cell_CAM.value = "CSS MHz Free"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 4)
                self.cell_CAM.value = "AP MHz Free"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 5)
                self.cell_CAM.value = "CSS MHz Used"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 6)
                self.cell_CAM.value = "AP MHz Used"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 7)
                self.cell_CAM.value = "CSS MHz Used By Individual Case"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 8)
                self.cell_CAM.value = "AP MHz Used By Individual Case"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 9)
                self.cell_CAM.value = "Voice Test Result"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 10)
                self.cell_CAM.value = "Average MHz (CSS)"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 11)
                self.cell_CAM.value = "Average MHz (APP)"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 12)
                self.cell_CAM.value = "Line Number(s) Used"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 13)
                self.cell_CAM.value = "RTP Mode"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 14)
                self.cell_CAM.value = "RTCP-XR"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 15)
                self.cell_CAM.value = "VAD"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 16)
                self.cell_CAM.value = "SRTP"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 17)
                self.cell_CAM.value = "Redundant RTP"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 18)
                self.cell_CAM.value = "Symmetric RTP"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 19)
                self.cell_CAM.value = "RTCP Feedback"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 20)
                self.cell_CAM.value = "RTCP MUX"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.workbook.save(self.report)
# For looper with DDR usage data
            elif self.app == 2:
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM)
                self.cell_CAM.value = self.suite_info
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.rows_CAM += 1
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM)
                self.cell_CAM.value = "Scenario"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 1)
                self.cell_CAM.value = "Looper"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 2)
                self.cell_CAM.value = "CSS Term CPU Free %"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 3)
                self.cell_CAM.value = "CSS MHz Free"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 4)
                self.cell_CAM.value = "AP MHz Free"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 5)
                self.cell_CAM.value = "CSS MHz Used"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 6)
                self.cell_CAM.value = "AP MHz Used"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 7)
                self.cell_CAM.value = "CSS MHz Used By Individual Case"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 8)
                self.cell_CAM.value = "AP MHz Used By Individual Case"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 9)
                self.cell_CAM.value = "DDR Usage (in %)"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 10)
                self.cell_CAM.value = "DDR Usage (in MBPS)"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 11)
                self.cell_CAM.value = "Voice Test Result"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 12)
                self.cell_CAM.value = "Average MHz (CSS)"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 13)
                self.cell_CAM.value = "Average MHz (APP)"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 14)
                self.cell_CAM.value = "Line Number(s) Used"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 15)
                self.cell_CAM.value = "RTP Mode"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 16)
                self.cell_CAM.value = "RTCP-XR"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 17)
                self.cell_CAM.value = "VAD"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 18)
                self.cell_CAM.value = "SRTP"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 19)
                self.cell_CAM.value = "Redundant RTP"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 20)
                self.cell_CAM.value = "Symmetric RTP"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 21)
                self.cell_CAM.value = "RTCP Feedback"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.cell_CAM = self.worksheet_CAM.cell(row=self.rows_CAM, column=self.cols_CAM + 22)
                self.cell_CAM.value = "RTCP MUX"
                self.cell_CAM.font = self.bold_font
                self.cell_CAM.fill = self.bgc
                self.workbook.save(self.report)

            self.rows_CAM += 1

        if self.java:
            self.worksheet_JSM = self.workbook.create_sheet()
            self.worksheet_JSM.title = 'JAVA_Score'
            for sheet in self.workbook:
                if sheet.title == 'Sheet':
                    self.workbook.remove_sheet(sheet)
            self.cell_JSM = self.worksheet_JSM.cell(row=self.rows_JSM, column=self.cols_JSM)
            self.cell_JSM.value = self.suite_info
            self.cell_JSM.font = self.bold_font
            self.cell_JSM.fill = self.bgc
            self.rows_JSM += 1
            self.cell_JSM = self.worksheet_JSM.cell(row=self.rows_JSM, column=self.cols_JSM)
            self.cell_JSM.value = "Scenario"
            self.cell_JSM.font = self.bold_font
            self.cell_JSM.fill = self.bgc
            self.cell_JSM = self.worksheet_JSM.cell(row=self.rows_JSM, column=self.cols_JSM + 1)
            self.cell_JSM.value = "Overall JAVA Score"
            self.cell_JSM.font = self.bold_font
            self.cell_JSM.fill = self.bgc
            self.workbook.save(self.report)

            self.rows_JSM += 1

        if self.df_dok:
            self.scenario_DBM = ['Anti-aliased Text','Anti-aliased Text (blend)','Fill Rectangle',
                                 'Fill Rectangle (blend)','Fill Rectangles [10]','Fill Rectangles [10] (blend)','Fill Triangles',
                                 'Fill Triangles (blend)','Draw Rectangle','Draw Rectangle (blend)','Draw Lines [10]','Draw Lines [10] (blend)',
                                 'Fill Spans','Fill Spans (blend)','Fill Trapezoids [10]','Blit','Blit 90','Blit 180','Blit 270','Flip Horizontal',
                                 'Flip Vertical','Blit colorkeyed','Blit destination colorkeyed','Blit with format conversion','Blit with colorizing',
                                 'Blit with mask','Blit from 32bit (blend)','Blit from 32bit (blend) with colorizing','Blit from 8bit palette',
                                 'Blit from 8bit palette (blend)','Blit SrcOver (premultiplied source)','Blit SrcOver (premultiply source)',
                                 'Stretch Blit','Stretch Blit colorkeyed','Stretch Blit YCbCr','Stretch Blit indexed','Stretch Flip Horizontal',
                                 'Stretch Flip Vertical','(128x128 RGB16)','(128x128 RGB16)','(128x128 RGB16)','(128x128 RGB16)']
            self.ldsl=len(self.scenario_DBM)
            self.worksheet_DBM = self.workbook.create_sheet()
            self.worksheet_DBM.title = 'DF_DOK'
            for sheet in self.workbook:
                if sheet.title == 'Sheet':
                    self.workbook.remove_sheet(sheet)
            self.cell_DBM = self.worksheet_DBM.cell(row=self.rows_DBM, column=self.cols_DBM)
            self.cell_DBM.value = self.suite_info
            self.cell_DBM.font = self.bold_font
            self.cell_DBM.fill = self.bgc
            self.rows_DBM += 1
            self.cell_DBM = self.worksheet_DBM.cell(row=self.rows_DBM, column=self.cols_DBM)
            self.cell_DBM.value = "Parameter"
            self.cell_DBM.font = self.bold_font
            self.cell_DBM.fill = self.bgc
            self.cell_DBM = self.worksheet_DBM.cell(row=self.rows_DBM, column=self.cols_DBM + 1)
            self.cell_DBM.value = "Kchars/Sec or Mpixels/Sec (Rate)"
            self.cell_DBM.font = self.bold_font
            self.cell_DBM.fill = self.bgc
            self.cell_DBM = self.worksheet_DBM.cell(row=self.rows_DBM, column=self.cols_DBM + 2)
            self.cell_DBM.value = "CPU Usage (in %)"
            self.cell_DBM.font = self.bold_font
            self.cell_DBM.fill = self.bgc
            self.workbook.save(self.report)

            self.rows_DBM += 1

        if self.bmr:
            self.worksheet_BMR = self.workbook.create_sheet()
            self.worksheet_BMR.title = 'Benchmarking_Score'
            for sheet in self.workbook:
                if sheet.title == 'Sheet':
                    self.workbook.remove_sheet(sheet)
            self.cell_BMR = self.worksheet_BMR.cell(row=self.rows_BMR, column=self.cols_BMR)
            self.cell_BMR.value = self.suite_info
            self.cell_BMR.font = self.bold_font
            self.cell_BMR.fill = self.bgc
            self.rows_BMR += 1
            self.cell_BMR = self.worksheet_BMR.cell(row=self.rows_BMR, column=self.cols_BMR)
            self.cell_BMR.value = "Scenario"
            self.cell_BMR.font = self.bold_font
            self.cell_BMR.fill = self.bgc
            self.cell_BMR = self.worksheet_BMR.cell(row=self.rows_BMR, column=self.cols_BMR + 1)
            self.cell_BMR.value = "Overall Benchmarking Score"
            self.cell_BMR.font = self.bold_font
            self.cell_BMR.fill = self.bgc
            self.workbook.save(self.report)
            self.rows_BMR += 1


    def update_perfReport(self, ws, ws_row, ws_col, cval):
        if self.DEBUG == 1:
            print "Updating performance report values...."
            print "Row No. : ", ws_row
            print "Column No. : ", ws_col
            print "Column No. : ", ws_col
            print "Value : ", cval
        if ws == 'RAM_Usage_Information':
            self.cell_RUI = self.worksheet_RUI.cell(row=ws_row, column=ws_col)
            if ws_col == 2:
                self.cell_RUI.value = cval
                self.cell_RUI.number_format = self.num_format
            else:
                self.cell_RUI.value = cval
            self.workbook.save(self.report)
        elif ws == 'CSS-AP_MHz_Measurement':
            self.cell_CAM = self.worksheet_CAM.cell(row=ws_row, column=ws_col)
            if ws_col >= 4:
                self.cell_CAM.value = cval
                self.cell_CAM.number_format = self.num_format
            elif ws_col == 2 or ws_col == 3:
                self.cell_CAM.value = float(cval)
            else:
                self.cell_CAM.value = cval
            self.workbook.save(self.report)
        elif ws == 'JAVA_Score':
            self.cell_JSM = self.worksheet_JSM.cell(row=ws_row, column=ws_col)
            if ws_col == 2:
                self.cell_JSM.value = int(cval)
            else:
                self.cell_JSM.value = cval
            self.workbook.save(self.report)
        elif ws == 'DF_DOK':
            self.cell_DBM = self.worksheet_DBM.cell(row=ws_row, column=ws_col)
            if ws_col >= 2:
                self.cell_DBM.value = cval
                self.cell_DBM.number_format = self.num_format
            else:
                self.cell_MSM.value = cval
            self.workbook.save(self.report)
        elif ws == 'Benchmarking_Score':
            self.cell_BMR = self.worksheet_BMR.cell(row=ws_row, column=ws_col)
            if ws_col == 2:
                self.cell_BMR.value = int(cval)
            else:
                self.cell_BMR.value = cval
            self.workbook.save(self.report)


    def voice_result_update(self):
        if self.app or self.css:
            if self.scenario not in ["Idle", "CSS+APP_DSP Load", "callManager Load", "WBHF Load"]:
                if self.app and self.ltype == 1:
                    self.update_perfReport('CSS-AP_MHz_Measurement', self.rows_CAM - 1, self.cols_CAM + 9, self.curr_voice_result)
                elif self.app and self.ltype == 2:
                    self.update_perfReport('CSS-AP_MHz_Measurement', self.rows_CAM - 1, self.cols_CAM + 11, self.curr_voice_result)
            if " Load" in self.scenario:
                if self.app and self.ltype == 1:
                    self.update_perfReport('CSS-AP_MHz_Measurement', self.rows_CAM - 1, self.cols_CAM + 9, "N/A")
                elif self.app and self.ltype == 2:
                    self.update_perfReport('CSS-AP_MHz_Measurement', self.rows_CAM - 1, self.cols_CAM + 11, "N/A")


    def calcMHzAvg(self):
        temp = []
        total_instances = 0
        sum_css_mhz = 0
        sum_app_mhz = 0
        book = openpyxl.load_workbook(self.report)
        sheet = book.active
        row_no = 1
        for i in sheet.rows:
            scenario_type = i[0].value
            if scenario_type:
                if re.search('Call',scenario_type):
                    if temp != [] and scenario_type[0] < temp[-1]:
                        average_css_mhz = sum_css_mhz/total_instances
                        average_app_mhz = sum_app_mhz/total_instances
                        sheet.cell(row = row_no - 1, column = 11).font = Font(bold=True)
                        sheet.cell(row = row_no - 1, column = 12).font = Font(bold=True)
                        sheet.cell(row = row_no - 1, column = 11).value = average_css_mhz
                        sheet.cell(row = row_no - 1, column = 12).value = average_app_mhz
                        sum_css_mhz = 0
                        sum_app_mhz = 0
                        total_instances = 0
                        temp = []
                    if scenario_type[0] == '1':
                        temp.append(scenario_type[0])
                        sum_css_mhz = sum_css_mhz + int(sheet.cell(row = row_no, column = 8).value)
                        sum_app_mhz = sum_app_mhz + int(sheet.cell(row = row_no, column = 9).value)
                        total_instances = total_instances + 1
                    else:
                        temp.append(scenario_type[0])
                        sum_css_mhz = sum_css_mhz + int(sheet.cell(row = row_no, column = 8).value)
                        sum_app_mhz = sum_app_mhz + int(sheet.cell(row = row_no, column = 9).value)
                        total_instances = total_instances + (int(scenario_type[0]) - 1)

            row_no = row_no + 1
        average_css_mhz = sum_css_mhz/total_instances
        average_app_mhz = sum_app_mhz/total_instances
        sheet.cell(row = row_no - 1, column = 11).font = Font(bold=True)
        sheet.cell(row = row_no - 1, column = 12).font = Font(bold=True)
        sheet.cell(row = row_no - 1, column = 11).value = average_css_mhz
        sheet.cell(row = row_no - 1, column = 12).value = average_app_mhz
        book.save(self.report)
        book.close()
# mhz_average('C:\MANU\\test\DVF101_SDK280-cand1-SPLIT-NDA_MHz.xlsx')



if __name__ == '__main__':
    print "Capturing...."
