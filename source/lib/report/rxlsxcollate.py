


########################################################################################################################
########################################################################################################################
## File              :: vaca_rxlsx_write_routines.py
## Description       :: VoIP Automation Common API : Collate and compare with previous results.
## Developer         :: Sreekanth S
## Version           :: v1.0
## Release Date      :: 11/04/2019
## Changes made      :: Initial version.
## Changes made Date :: 11/04/2019
## Changes made by   :: Sreekanth S
########################################################################################################################
########################################################################################################################



import time
import openpyxl
import re
from copy import copy
from openpyxl.styles import Font


# TODO: Compare and add usable fucntions or statements to rxlsxwrite.py and remove or improve this file.


class eXLSXreport:
    def __init__(self):
        self.path = ""
        self.report = ""
        self.ram = 0
        self.css = 0
        self.app = 0
        self.java = 0
        self.dfdok = 0
        self.bmr = 0
        self.rows_RUI = 1
        self.cols_RUI = 1
        self.rows_CAM = 1
        self.cols_CAM = 1
        self.rows_JSM = 1
        self.cols_JSM = 1
        self.rows_DBM = 1
        self.cols_DBM = 1
        self.rows_BMR = 1
        self.cols_BMR = 1
        self.workbook = openpyxl.Workbook()
        self.worksheet_RUI = ""
        self.worksheet_CAM = ""
        self.worksheet_JSM = ""
        self.worksheet_DBM = ""
        self.worksheet_BMR = ""
        self.cell_RUI = ""
        self.cell_CAM = ""
        self.cell_JSM = ""
        self.cell_DBM = ""
        self.cell_BMR = ""
        self.num_format = '#0.000000'
        self.bold_font = openpyxl.styles.Font(bold=True)


    def calcMHzAvg(self):
        temp = []
        total_instances = 0
        sum_css_mhz = 0
        sum_app_mhz = 0
        book = openpyxl.load_workbook(self.report)
        sheet = book.active
        row_no = 1
        for i in sheet.rows:
            scenario_type = i[0].value
            if scenario_type:
                if re.search('Call',scenario_type):
                    if temp != [] and scenario_type[0] < temp[-1]:
                        average_css_mhz = sum_css_mhz/total_instances
                        average_app_mhz = sum_app_mhz/total_instances
                        sheet.cell(row = row_no - 1, column = 11).font = Font(size=17, bold=True)
                        sheet.cell(row = row_no - 1, column = 12).font = Font(size=17, bold=True)
                        sheet.cell(row = row_no - 1, column = 11).value = average_css_mhz
                        sheet.cell(row = row_no - 1, column = 12).value = average_app_mhz
                        sum_css_mhz = 0
                        sum_app_mhz = 0
                        total_instances = 0
                        temp = []
                    if scenario_type[0] == '1':
                        temp.append(scenario_type[0])
                        sum_css_mhz = sum_css_mhz + int(sheet.cell(row = row_no, column = 8).value)
                        sum_app_mhz = sum_app_mhz + int(sheet.cell(row = row_no, column = 9).value)
                        total_instances = total_instances + 1
                    else:
                        temp.append(scenario_type[0])
                        sum_css_mhz = sum_css_mhz + int(sheet.cell(row = row_no, column = 8).value)
                        sum_app_mhz = sum_app_mhz + int(sheet.cell(row = row_no, column = 9).value)
                        total_instances = total_instances + (int(scenario_type[0]) - 1)

            row_no = row_no + 1
        average_css_mhz = sum_css_mhz/total_instances
        average_app_mhz = sum_app_mhz/total_instances
        sheet.cell(row = row_no - 1, column = 11).font = Font(size=17, bold=True)
        sheet.cell(row = row_no - 1, column = 12).font = Font(size=17, bold=True)
        sheet.cell(row = row_no - 1, column = 11).value = average_css_mhz
        sheet.cell(row = row_no - 1, column = 12).value = average_app_mhz
        book.save(self.report)
        book.close()


    def update_Voipunits_RTP_VAD(xlsxfile,configfile):
        temp = []
        total_instances = 0
        sum_css_mhz = 0
        sum_app_mhz = 0
        book = openpyxl.load_workbook(xlsxfile)
        sheet = book.active
        sheet.cell(2, 11, 'CSS-MHz-Average')._style = copy(sheet.cell(2, 10)._style)
        sheet.cell(2, 12, 'APP-MHz-Average')._style = copy(sheet.cell(2, 10)._style)
        sheet.cell(2, 13, 'Voip Units')._style = copy(sheet.cell(2, 10)._style)
        sheet.cell(2, 14, 'RTP/SRTP')._style = copy(sheet.cell(2, 10)._style)
        sheet.cell(2, 15, 'VAD')._style = copy(sheet.cell(2, 10)._style)
        row_no = 1
        fd = open(configfile, 'r')
        lines = fd.readline()
        for i in sheet.rows:
            scenario_type = i[0].value
            if scenario_type:
                if re.search('Call', scenario_type):
                    while not lines[0].isdigit():
                         lines = fd.readline()
                    if scenario_type[0] == '1':
                        lines_list = lines.split(',')
                        sheet.cell(row_no, 13).value = 'line No - ' + lines_list[2]
                        if int(lines_list[4]):
                            sheet.cell(row_no, 14).value = 'SRTP'
                        else:
                            sheet.cell(row_no, 14).value = 'RTP'
                        sheet.cell(row_no, 15).value = lines_list[6]
                        lines = fd.readline()
                    else:
                        k = 1
                        line_nos = 'Line No - '
                        lines_list = lines.split(',')
                        while k < int(lines_list[0][-1]):
                            line_nos = line_nos + lines_list[k*2] + ','
                            k = k + 1
                        sheet.cell(row_no, 13).value = line_nos.strip(',')
                        if int(lines_list[(int(lines_list[0][-1]) * 2) + (int(lines_list[0][-1]) - 2)]):
                            sheet.cell(row_no, 14).value = 'SRTP'
                        else:
                            sheet.cell(row_no, 14).value = 'RTP'
                        sheet.cell(row_no, 15).value = lines_list[(int(lines_list[0][-1]) * 2) + int(lines_list[0][-1])]
                        lines = fd.readline()
            row_no = row_no + 1
        fd.close()
        book.save(xlsxfile)
        book.close()


# update_Voipunits_RTP_VAD('C:\\MANU\\test\\DVF101_SDK280-cand1-SPLIT-NDA_MHz.xlsx', 'C:\\MANU\\test\\config_user.txt')
# mhz_average('C:\MANU\\test\DVF101_SDK280-cand1-SPLIT-NDA_MHz.xlsx')



if __name__ == '__main__':
    print "Capturing...."
